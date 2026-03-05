from __future__ import annotations

from datetime import date, datetime
import json
from math import ceil
import os
import re
import unicodedata
import uuid
from contextlib import contextmanager
from typing import Any, Generator, Protocol

import httpx
import psycopg2
from fastapi import FastAPI, HTTPException
from fastapi.concurrency import run_in_threadpool
from psycopg2.extensions import connection as PgConnection
from psycopg2.extras import execute_batch
from pydantic import BaseModel, Field
from sentence_transformers import SentenceTransformer

try:
    from llama_cpp import Llama
except Exception as exc:  # pragma: no cover - optional runtime dependency
    Llama = Any  # type: ignore[assignment]
    LLAMA_CPP_IMPORT_ERROR: Exception | None = exc
else:
    LLAMA_CPP_IMPORT_ERROR = None

try:
    from huggingface_hub import hf_hub_download
except Exception as exc:  # pragma: no cover - optional runtime dependency
    hf_hub_download = None
    HUGGINGFACE_IMPORT_ERROR: Exception | None = exc
else:
    HUGGINGFACE_IMPORT_ERROR = None

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - optional runtime dependency
    load_workbook = None
    OPENPYXL_IMPORT_ERROR: Exception | None = exc
else:
    OPENPYXL_IMPORT_ERROR = None


EMBEDDING_MODEL_NAME = "sentence-transformers/all-MiniLM-L6-v2"
EMBEDDING_DIMENSION = 384
REINDEX_BATCH_SIZE = 100
DEFAULT_TOP_K = 5

LLM_PROVIDER = os.getenv("LLM_PROVIDER", "ollama").strip().lower()
OLLAMA_URL = os.getenv("OLLAMA_URL", "http://127.0.0.1:11434").rstrip("/")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "phi3")
INGEST_EXCEL_BASE_DIR = os.path.abspath(
    os.path.normpath(
        os.getenv("INGEST_EXCEL_BASE_DIR", r"C:\laragon\www\gore-chatbot-api\storage\app\kb")
    )
)

LLM_MODEL_REPO = os.getenv("LLM_MODEL_REPO", "TheBloke/TinyLlama-1.1B-Chat-v1.0-GGUF")
LLM_MODEL_FILE = os.getenv("LLM_MODEL_FILE", "tinyllama-1.1b-chat-v1.0.Q4_K_M.gguf")
LLM_MODEL_PATH = os.getenv("LLM_MODEL_PATH")
LLM_MAX_TOKENS = int(os.getenv("LLM_MAX_TOKENS", "256"))
LLM_TEMPERATURE = float(os.getenv("LLM_TEMPERATURE", "0.2"))
LLM_CTX = int(os.getenv("LLM_CTX", "2048"))
LLM_THREADS = int(os.getenv("LLM_THREADS", "4"))

PG_HOST = os.getenv("PG_HOST", "127.0.0.1")
PG_PORT = int(os.getenv("PG_PORT", "5433"))
PG_DATABASE = os.getenv("PG_DATABASE", "SIA")
PG_USER = os.getenv("PG_USER", "postgres")
PG_PASSWORD = os.getenv("PG_PASSWORD", "123456")

embedding_model: SentenceTransformer | None = None
llm_model: Llama | None = None
llm_provider_instance: "LlmProvider | None" = None

NO_SUPPORT_ANSWER = "No encuentro sustento en la base cargada para responder eso."
SENSITIVE_DATA_PATTERNS = [
    re.compile(r"\b\d+\s*(d[ií]as?|horas?|mes(es)?|años?)\b", re.IGNORECASE),
    re.compile(r"h[aá]biles", re.IGNORECASE),
    re.compile(r"S\/\s*\d+", re.IGNORECASE),
    re.compile(r"\b\d+\s*%\b", re.IGNORECASE),
    re.compile(r"soles", re.IGNORECASE),
]
NODE_TITLE_TOKEN_PATTERN = re.compile(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9]{4,}")

NODE_SCORE_STRONG_THRESHOLD = 0.70
NODE_SCORE_MARGIN_THRESHOLD = 0.12
NODE_VERIFICATION_MIN_RATIO = 0.4
NODE_VERIFICATION_MIN_HITS = 1
NODE_TITLE_STOPWORDS = {
    "del",
    "de",
    "la",
    "los",
    "las",
    "para",
    "por",
    "con",
    "sin",
    "que",
    "una",
    "uno",
    "y",
    "el",
    "al",
    "procedimiento",
    "administrativo",
    "administrativa",
}

PROCEDURE_CHUNK_FIELDS = [
    "descripcion_procedimiento",
    "requisitos",
    "formularios",
    "canales_atencion",
    "pago_tramite",
    "modalidad_pago",
    "calificacion_procedimiento",
    "sedes_horarios_atencion",
    "unidad_org_presenta_documentacion",
    "unidad_org_responsable_aprobar",
    "consulta_procedimiento",
    "instancias_resolucion_recursos",
    "plazo",
    "base_legal",
]

ORG_UNIT_CHUNK_FIELDS = [
    "dependencia_superior",
    "funciones",
    "atribuciones",
    "canales_contacto",
    "base_legal",
]

DOCUMENT_HEADERS = [
    "document_key",
    "title",
    "doc_type",
    "source_url",
    "status",
    "version",
    "issued_date",
    "entity",
    "notes",
]

TUPA_HEADERS = [
    "document_key",
    "procedure_code",
    "procedure_title",
    "descripcion_procedimiento",
    "requisitos",
    "formularios",
    "canales_atencion",
    "pago_tramite",
    "modalidad_pago",
    "plazo",
    "calificacion_procedimiento",
    "sedes_horarios_atencion",
    "unidad_org_presenta_documentacion",
    "unidad_org_responsable_aprobar",
    "consulta_procedimiento",
    "instancias_resolucion_recursos",
    "base_legal",
    "pdf_page_start",
    "pdf_page_end",
    "status",
]

ROF_HEADERS = [
    "document_key",
    "org_code",
    "org_title",
    "dependencia_superior",
    "funciones",
    "atribuciones",
    "canales_contacto",
    "base_legal",
    "pdf_page_start",
    "pdf_page_end",
    "status",
    "tags",
]

SENSITIVE_FIELD_PRIORITY_RULES_TUPA: list[tuple[set[str], list[str]]] = [
    # descripcion_procedimiento
    (
        {
            "descripcion",
            "descripción",
            "trata",
            "consiste",
            "de que trata",
            "sobre que es",
            "objetivo",
            "finalidad",
            "que es el tramite",
            "qué es el trámite",
        },
        ["descripcion_procedimiento"],
    ),
    # requisitos
    (
        {
            "requisito",
            "requisitos",
            "documento",
            "documentos",
            "adjuntar",
            "presentar",
            "presentacion",
            "presentación",
            "que necesito",
            "qué necesito",
            "que debo presentar",
            "qué debo presentar",
            "anexo",
            "anexos",
        },
        ["requisitos"],
    ),
    # formularios
    (
        {
            "formulario",
            "formularios",
            "formato",
            "formatos",
            "fut",
            "solicitud",
            "modelo de solicitud",
            "plantilla",
            "descargar formulario",
        },
        ["formularios"],
    ),
    # canales_atencion
    (
        {
            "canal",
            "canales",
            "canal de atencion",
            "canal de atención",
            "virtual",
            "presencial",
            "mesa de partes virtual",
            "tramite en linea",
            "trámite en línea",
            "web",
            "telefono",
            "teléfono",
            "correo",
            "email",
        },
        ["canales_atencion"],
    ),
    # pago_tramite (cuanto cuesta)
    (
        {
            "costo",
            "costos",
            "monto",
            "precio",
            "tarifa",
            "importe",
            "valor",
            "derecho de tramite",
            "derecho de trámite",
            "cuanto cuesta",
            "cuánto cuesta",
            "cuanto debo pagar",
            "cuánto debo pagar",
            "pago del tramite",
            "pago del trámite",
            "cuanto es",
            "cuánto es",
        },
        ["pago_tramite"],
    ),
    # modalidad_pago (como se paga)
    (
        {
            "modalidad de pago",
            "formas de pago",
            "forma de pago",
            "medio de pago",
            "medios de pago",
            "como pagar",
            "cómo pagar",
            "donde pagar",
            "dónde pagar",
            "efectivo",
            "transferencia",
            "deposito",
            "depósito",
            "yape",
            "plin",
            "tarjeta",
            "credito",
            "crédito",
            "debito",
            "débito",
            "cheque",
            "banco",
            "cuenta bancaria",
            "pago online",
            "qr",
            "pos",
        },
        ["modalidad_pago"],
    ),
    # plazo
    (
        {
            "plazo",
            "plazos",
            "dias",
            "días",
            "dia",
            "día",
            "dias habiles",
            "días hábiles",
            "tiempo",
            "demora",
            "cuanto demora",
            "cuánto demora",
            "en cuanto tiempo",
            "en cuánto tiempo",
            "cuando sale",
            "cuándo sale",
        },
        ["plazo"],
    ),
    # calificacion_procedimiento
    (
        {
            "calificacion",
            "calificación",
            "evaluacion previa",
            "evaluación previa",
            "aprobacion automatica",
            "aprobación automática",
            "silencio administrativo",
            "silencio positivo",
            "silencio negativo",
            "tipo de procedimiento",
        },
        ["calificacion_procedimiento"],
    ),
    # sedes_horarios_atencion
    (
        {
            "sede",
            "sedes",
            "oficina",
            "oficinas",
            "direccion",
            "dirección",
            "ubicacion",
            "ubicación",
            "lugar",
            "donde atienden",
            "dónde atienden",
            "horario",
            "horarios",
            "hora de atencion",
            "hora de atención",
            "atencion al publico",
            "atención al público",
        },
        ["sedes_horarios_atencion"],
    ),
    # unidad_org_presenta_documentacion
    (
        {
            "donde presento",
            "dónde presento",
            "donde se presenta",
            "dónde se presenta",
            "unidad presenta documentacion",
            "unidad presenta documentación",
            "mesa de partes",
            "recepcion de documentos",
            "recepción de documentos",
            "oficina receptora",
        },
        ["unidad_org_presenta_documentacion"],
    ),
    # unidad_org_responsable_aprobar
    (
        {
            "quien aprueba",
            "quién aprueba",
            "unidad responsable",
            "organo responsable",
            "órgano responsable",
            "area responsable",
            "área responsable",
            "autoridad competente",
            "gerencia responsable",
        },
        ["unidad_org_responsable_aprobar"],
    ),
    # consulta_procedimiento
    (
        {
            "consulta",
            "consultar",
            "seguimiento",
            "estado del tramite",
            "estado del trámite",
            "como consultar",
            "cómo consultar",
            "telefono de consulta",
            "teléfono de consulta",
            "correo de consulta",
            "contacto",
            "informes",
        },
        ["consulta_procedimiento"],
    ),
    # instancias_resolucion_recursos
    (
        {
            "instancia",
            "instancias",
            "recurso",
            "recursos",
            "apelacion",
            "apelación",
            "reconsideracion",
            "reconsideración",
            "impugnacion",
            "impugnación",
            "segunda instancia",
            "resolucion de recursos",
            "resolución de recursos",
        },
        ["instancias_resolucion_recursos"],
    ),
    # base_legal
    (
        {
            "base legal",
            "norma",
            "normativa",
            "ley",
            "decreto",
            "reglamento",
            "sustento legal",
            "marco legal",
            "fundamento legal",
        },
        ["base_legal"],
    ),
]

SENSITIVE_FIELD_PRIORITY_RULES_ROF: list[tuple[set[str], list[str]]] = [
    ({"funcion", "funciones", "rol", "roles", "competencia", "competencias"}, ["funciones"]),
    ({"atribucion", "atribuciones", "facultad", "facultades"}, ["atribuciones"]),
    (
        {"dependencia", "depende", "superior", "jefatura", "jefe", "gerencia"},
        ["dependencia_superior"],
    ),
    ({"canal", "canales", "contacto", "telefono", "correo", "email"}, ["canales_contacto"]),
    ({"norma", "normativa", "ley", "base legal"}, ["base_legal"]),
]

SENSITIVE_FIELD_PRIORITY_RULES_DEFAULT: list[tuple[set[str], list[str]]] = []

SENSITIVE_FIELD_PRIORITY_RULES_BY_DOCUMENT_KEY: dict[str, list[tuple[set[str], list[str]]]] = {
    "tupa": SENSITIVE_FIELD_PRIORITY_RULES_TUPA,
    "rof": SENSITIVE_FIELD_PRIORITY_RULES_ROF,
}

SENSITIVE_FIELD_PRIORITY_RULES_BY_NODE_TYPE: dict[str, list[tuple[set[str], list[str]]]] = {
    "procedure": SENSITIVE_FIELD_PRIORITY_RULES_TUPA,
    "org_unit": SENSITIVE_FIELD_PRIORITY_RULES_ROF,
}


class RetrieveFilters(BaseModel):
    procedure_title: str | None = Field(default=None, min_length=1)
    document_id: str | None = Field(default=None, min_length=1)
    node_id: str | None = Field(default=None, min_length=1)


class RetrieveRequest(BaseModel):
    query: str = Field(..., min_length=1)
    top_k: int = Field(default=DEFAULT_TOP_K, ge=1, le=20)
    filters: RetrieveFilters | None = None


class RetrievedDocument(BaseModel):
    title: str
    type: str
    url: str | None = None


class RetrieveResult(BaseModel):
    chunk_id: str
    content: str
    document: RetrievedDocument


class RetrieveResponse(BaseModel):
    results: list[RetrieveResult]


class NodeCandidate(BaseModel):
    node_id: str
    title: str
    code: str | None = None
    node_type: str
    document_id: str
    score: float


class ResolveNodesRequest(BaseModel):
    query: str = Field(..., min_length=1)
    document_id: str | None = Field(default=None, min_length=1)
    node_type: str | None = Field(default=None, min_length=1, max_length=50)
    limit: int = Field(default=5, ge=1, le=10)


class ResolveNodesResponse(BaseModel):
    strong_match: bool
    selected_node_id: str | None = None
    candidates: list[NodeCandidate]


class ChatRequest(BaseModel):
    message: str = Field(..., min_length=1)
    top_k: int = Field(default=DEFAULT_TOP_K, ge=1, le=20)
    selected_node_id: str | None = Field(default=None, min_length=1)
    document_id: str | None = Field(default=None, min_length=1)
    conversation_id: str | None = Field(default=None, min_length=1)


class ChatSource(BaseModel):
    title: str
    type: str
    reference: str
    url: str | None = None


class ChatResponse(BaseModel):
    answer: str
    sources: list[ChatSource]

class IngestExcelRequest(BaseModel):
    path: str
    mode: str = Field(default="upsert")
    embed: bool = Field(default=True)


class IngestExcelResponse(BaseModel):
    documents_upserted: int
    nodes_upserted: int
    chunks_upserted: int
    embedded: int
    errors: list[str]


class ReindexResponse(BaseModel):
    processed: int
    remaining: int


class ReindexNodeRequest(BaseModel):
    node_id: str = Field(..., min_length=1)
    force: bool = True


class ReindexNodeResponse(BaseModel):
    node_id: str
    processed: int
    remaining: int


app = FastAPI(
    title="GORE Chatbot AI Service",
    version="0.1.0",
    description="Stub service for retrieve() used by Laravel phase 2.",
)


class LlmProvider(Protocol):
    def generate(self, prompt: str, max_tokens: int, temperature: float) -> str:
        ...


class OllamaProvider:
    def __init__(self, base_url: str, model: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.model = model

    def generate(self, prompt: str, max_tokens: int, temperature: float) -> str:
        endpoint = f"{self.base_url}/api/generate"
        payload = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
            "options": {
                "temperature": temperature,
                "num_predict": max_tokens,
            },
        }

        try:
            response = httpx.post(endpoint, json=payload, timeout=120.0)
        except httpx.RequestError as exc:
            raise RuntimeError(
                f"Ollama no responde en {self.base_url}. "
                "Verifique que el servicio este activo."
            ) from exc

        if response.status_code >= 400:
            raise RuntimeError(
                f"Ollama devolvio HTTP {response.status_code}: {response.text[:300]}"
            )

        data = response.json()
        answer = str(data.get("response", "")).strip()

        if answer == "":
            raise RuntimeError("Ollama devolvio una respuesta vacia.")

        return answer


class LlamaCppProvider:
    def generate(self, prompt: str, max_tokens: int, temperature: float) -> str:
        llm = get_llm()
        completion = llm(
            prompt=prompt,
            max_tokens=max_tokens,
            temperature=temperature,
            stop=["\nUsuario:", "\nSistema:"],
        )

        choices = completion.get("choices", []) if isinstance(completion, dict) else []
        if not choices:
            raise RuntimeError("llama.cpp no devolvio elecciones.")

        text = str(choices[0].get("text", "")).strip()
        if text == "":
            raise RuntimeError("llama.cpp devolvio una respuesta vacia.")

        return text


def get_db_connection() -> PgConnection:
    return psycopg2.connect(
        host=PG_HOST,
        port=PG_PORT,
        dbname=PG_DATABASE,
        user=PG_USER,
        password=PG_PASSWORD,
    )


@contextmanager
def db_connection() -> Generator[PgConnection, None, None]:
    connection = get_db_connection()
    try:
        yield connection
    finally:
        connection.close()


def get_embedding_model() -> SentenceTransformer:
    global embedding_model

    if embedding_model is None:
        embedding_model = SentenceTransformer(EMBEDDING_MODEL_NAME)

    return embedding_model


def resolve_llm_model_path() -> str:
    if LLM_MODEL_PATH and os.path.isfile(LLM_MODEL_PATH):
        return LLM_MODEL_PATH

    if LLM_MODEL_REPO and LLM_MODEL_FILE:
        if hf_hub_download is None:
            raise RuntimeError(
                f"huggingface_hub is not available: {HUGGINGFACE_IMPORT_ERROR}"
            )

        return hf_hub_download(repo_id=LLM_MODEL_REPO, filename=LLM_MODEL_FILE)

    raise RuntimeError(
        "LLM model path is not configured. Set LLM_MODEL_PATH or LLM_MODEL_REPO + LLM_MODEL_FILE."
    )


def get_llm() -> Llama:
    global llm_model

    if llm_model is None:
        if LLAMA_CPP_IMPORT_ERROR is not None:
            raise RuntimeError(f"llama-cpp-python is not available: {LLAMA_CPP_IMPORT_ERROR}")

        model_path = resolve_llm_model_path()
        llm_model = Llama(
            model_path=model_path,
            n_ctx=LLM_CTX,
            n_threads=LLM_THREADS,
            n_gpu_layers=0,
            verbose=False,
        )

    return llm_model


def get_llm_provider() -> LlmProvider:
    global llm_provider_instance

    if llm_provider_instance is None:
        if LLM_PROVIDER == "ollama":
            llm_provider_instance = OllamaProvider(
                base_url=OLLAMA_URL,
                model=OLLAMA_MODEL,
            )
        elif LLM_PROVIDER == "llamacpp":
            llm_provider_instance = LlamaCppProvider()
        else:
            raise RuntimeError(
                f"Unsupported LLM_PROVIDER '{LLM_PROVIDER}'. Use 'ollama' or 'llamacpp'."
            )

    return llm_provider_instance


def to_vector_literal(values: list[float]) -> str:
    formatted = ",".join(f"{float(value):.8f}" for value in values)
    return f"[{formatted}]"


def normalize_text(text: str) -> str:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"[ \t]+", " ", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    normalized = re.sub(r"[ \t]*\n[ \t]*", "\n", normalized)
    normalized = re.sub(r"\s{2,}", " ", normalized)
    return normalized.strip()


def validate_ingest_excel_path(input_path: str) -> str:
    if input_path.strip() == "":
        raise ValueError("El path del Excel es obligatorio.")

    resolved = os.path.abspath(os.path.realpath(os.path.normpath(input_path.strip())))
    base_resolved = os.path.abspath(os.path.realpath(os.path.normpath(INGEST_EXCEL_BASE_DIR)))
    if os.name == "nt":
        resolved_cmp = resolved.casefold()
        base_cmp = base_resolved.casefold()
    else:
        resolved_cmp = resolved
        base_cmp = base_resolved

    if not resolved.lower().endswith(".xlsx"):
        raise ValueError(
            "El archivo debe tener extension .xlsx. "
            f"Path recibido: {input_path}. Path normalizado: {resolved}"
        )

    if not os.path.isfile(resolved):
        raise ValueError(
            f"El archivo no existe: {resolved}. "
            f"Path recibido: {input_path}. Path normalizado: {resolved}"
        )

    try:
        in_allowed_base = os.path.commonpath([resolved_cmp, base_cmp]) == base_cmp
    except ValueError as exc:
        raise ValueError(
            "Path invalido para ingest/excel. "
            f"Path recibido: {input_path}. Path normalizado: {resolved}"
        ) from exc

    if not in_allowed_base:
        raise ValueError(
            "Path fuera del directorio permitido. "
            f"Solo se permite bajo: {base_resolved}. "
            f"Path recibido: {input_path}. Path normalizado: {resolved}"
        )

    return resolved


def excel_value_to_str(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        normalized = normalize_text(value)
        return normalized if normalized != "" else None

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    normalized = normalize_text(str(value))
    return normalized if normalized != "" else None


def excel_value_to_int(value: Any) -> int | None:
    if value is None:
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value)

    text = excel_value_to_str(value)
    if text is None:
        return None

    try:
        return int(float(text))
    except ValueError:
        return None


def normalize_excel_header(value: Any) -> str:
    text = excel_value_to_str(value)
    if text is None:
        return ""
    normalized = unicodedata.normalize("NFKD", text.strip().lower())
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return normalized


def get_sheet_headers(workbook: Any, sheet_name: str) -> list[str]:
    if sheet_name not in workbook.sheetnames:
        return []

    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True, max_row=1))
    if not rows:
        return []

    headers: list[str] = []
    for value in rows[0]:
        header = normalize_excel_header(value)
        if header != "":
            headers.append(header)
    return headers


def validate_sheet_headers(
    workbook: Any,
    sheet_name: str,
    expected_headers: list[str],
) -> str | None:
    headers = get_sheet_headers(workbook, sheet_name)
    if not headers:
        return f"Hoja {sheet_name} vacia o inexistente."

    missing = [header for header in expected_headers if header not in headers]
    unexpected = [header for header in headers if header not in expected_headers]

    if missing or unexpected:
        missing_text = ", ".join(missing) if missing else "-"
        unexpected_text = ", ".join(unexpected) if unexpected else "-"
        return (
            f"Cabeceras invalidas en {sheet_name}. "
            f"Faltantes: [{missing_text}]. "
            f"No permitidas: [{unexpected_text}]."
        )

    return None


def validate_optional_sheet_headers(
    workbook: Any,
    sheet_name: str,
    expected_headers: list[str],
) -> str | None:
    if sheet_name not in workbook.sheetnames:
        return None

    headers = get_sheet_headers(workbook, sheet_name)
    # Optional sheet: if present but empty, skip validation and treat as no-data sheet.
    if not headers:
        return None

    missing = [header for header in expected_headers if header not in headers]
    unexpected = [header for header in headers if header not in expected_headers]

    if missing or unexpected:
        missing_text = ", ".join(missing) if missing else "-"
        unexpected_text = ", ".join(unexpected) if unexpected else "-"
        return (
            f"Cabeceras invalidas en {sheet_name}. "
            f"Faltantes: [{missing_text}]. "
            f"No permitidas: [{unexpected_text}]."
        )

    return None


def parse_tags(value: Any) -> list[str]:
    text = excel_value_to_str(value)
    if text is None:
        return []

    parts = re.split(r"[;,|]", text)
    tags: list[str] = []
    seen: set[str] = set()
    for part in parts:
        tag = normalize_text(part).strip().lower()
        if tag == "" or tag in seen:
            continue
        seen.add(tag)
        tags.append(tag)

    return tags


def read_sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []

    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_excel_header(header) for header in rows[0]]
    result: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(excel_value_to_str(cell) is not None for cell in row):
            continue

        record: dict[str, Any] = {"_row_number": row_number}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            record[header] = value

        result.append(record)

    return result


def embed_text(text: str) -> str:
    model = get_embedding_model()
    embedding = model.encode(
        text,
        convert_to_numpy=True,
        show_progress_bar=False,
        normalize_embeddings=True,
    )
    embedding_list = embedding.tolist()

    if len(embedding_list) != EMBEDDING_DIMENSION:
        raise ValueError(
            f"Invalid query embedding dimension: {len(embedding_list)} "
            f"(expected {EMBEDDING_DIMENSION})"
        )

    return to_vector_literal(embedding_list)


def retrieve_semantic(
    query: str,
    top_k: int,
    filters: RetrieveFilters | None = None,
    preferred_fields: list[str] | None = None,
) -> list[RetrieveResult]:
    query_vector = embed_text(query)

    base_query = """
        SELECT
            c.id,
            c.document_id,
            c.chunk_index,
            c.content,
            c.page,
            d.title,
            d.type,
            d.source_url
        FROM kb_chunks c
        JOIN kb_documents d ON c.document_id = d.id
        LEFT JOIN kb_nodes n ON c.node_id = n.id
        WHERE c.embedding IS NOT NULL
          AND d.status = 'published'
    """
    params: list[Any] = []

    if filters is not None:
        if filters.document_id:
            base_query += " AND c.document_id = %s::uuid"
            params.append(filters.document_id.strip())

        if filters.node_id:
            base_query += " AND c.node_id = %s::uuid AND n.status = 'published'"
            params.append(filters.node_id.strip())

        if filters.procedure_title:
            base_query += " AND c.metadata->>'procedure_title' = %s"
            params.append(filters.procedure_title.strip())

    if preferred_fields:
        base_query += """
            ORDER BY
                CASE WHEN c.metadata->>'field' = ANY(%s::text[]) THEN 0 ELSE 1 END,
                c.embedding <=> %s::vector
            LIMIT %s
        """
        params.extend([preferred_fields, query_vector, top_k])
    else:
        base_query += """
            ORDER BY c.embedding <=> %s::vector
            LIMIT %s
        """
        params.extend([query_vector, top_k])

    with db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(base_query, tuple(params))
            rows = cursor.fetchall()

    return [
        RetrieveResult(
            chunk_id=str(row[0]),
            content=str(row[3]),
            document=RetrievedDocument(
                title=str(row[5]),
                type=str(row[6]),
                url=(str(row[7]) if row[7] is not None else None),
            ),
        )
        for row in rows
    ]


def tokenize_for_node(text: str) -> set[str]:
    return {
        token.lower()
        for token in NODE_TITLE_TOKEN_PATTERN.findall(text)
        if token.lower() not in NODE_TITLE_STOPWORDS
    }


def score_lexical_node_match(query: str, title: str, code: str | None) -> float:
    query_clean = query.strip().lower()
    title_clean = title.strip().lower()
    if query_clean == "" or title_clean == "":
        return 0.0

    score = 0.0
    if query_clean in title_clean:
        score += 0.6

    query_tokens = tokenize_for_node(query_clean)
    title_tokens = tokenize_for_node(title_clean)
    if query_tokens:
        overlap = len(query_tokens.intersection(title_tokens)) / max(1, len(query_tokens))
        score += 0.35 * overlap

    if code and code.lower() in query_clean:
        score += 0.25

    return min(score, 1.0)


def fetch_published_node(node_id: str) -> dict[str, str | None] | None:
    with db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT n.id, n.document_id, n.node_type, n.title, n.code, d.document_key
                FROM kb_nodes n
                JOIN kb_documents d ON d.id = n.document_id
                WHERE n.id = %s::uuid
                  AND n.status = 'published'
                  AND d.status = 'published'
                LIMIT 1
                """,
                (node_id,),
            )
            row = cursor.fetchone()

    if row is None:
        return None

    return {
        "id": str(row[0]),
        "document_id": str(row[1]),
        "node_type": str(row[2]),
        "title": str(row[3]),
        "code": (str(row[4]) if row[4] is not None else None),
        "document_key": (str(row[5]) if row[5] is not None else None),
    }


def resolve_nodes_hybrid(
    query: str,
    document_id: str | None = None,
    node_type: str | None = None,
    limit: int = 5,
) -> ResolveNodesResponse:
    query_vector = embed_text(query)
    lexical_rows: list[tuple[Any, ...]] = []
    semantic_rows: list[tuple[Any, ...]] = []

    lexical_conditions = ["n.title ILIKE %s", "COALESCE(n.code, '') ILIKE %s"]
    lexical_params: list[Any] = [f"%{query.strip()}%", f"%{query.strip()}%"]
    for token in extract_search_terms(query, max_terms=8):
        lexical_conditions.append("n.title ILIKE %s")
        lexical_params.append(f"%{token}%")

    lexical_query = """
        SELECT n.id, n.document_id, n.node_type, n.title, n.code
        FROM kb_nodes n
        JOIN kb_documents d ON d.id = n.document_id
        WHERE n.status = 'published'
          AND d.status = 'published'
    """
    if document_id:
        lexical_query += " AND n.document_id = %s::uuid"
        lexical_params.append(document_id)
    if node_type:
        lexical_query += " AND n.node_type = %s"
        lexical_params.append(node_type)

    lexical_query += f" AND ({' OR '.join(lexical_conditions)})"
    lexical_query += " ORDER BY n.updated_at DESC LIMIT 120"

    semantic_query = """
        SELECT
            c.node_id,
            n.document_id,
            n.node_type,
            n.title,
            n.code,
            MIN(c.embedding <=> %s::vector) AS min_distance
        FROM kb_chunks c
        JOIN kb_nodes n ON c.node_id = n.id
        JOIN kb_documents d ON d.id = n.document_id
        WHERE c.embedding IS NOT NULL
          AND c.node_id IS NOT NULL
          AND n.status = 'published'
          AND d.status = 'published'
    """
    semantic_params: list[Any] = [query_vector]
    if document_id:
        semantic_query += " AND n.document_id = %s::uuid"
        semantic_params.append(document_id)
    if node_type:
        semantic_query += " AND n.node_type = %s"
        semantic_params.append(node_type)

    semantic_query += """
        GROUP BY c.node_id, n.document_id, n.node_type, n.title, n.code
        ORDER BY min_distance ASC
        LIMIT 40
    """

    with db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(lexical_query, tuple(lexical_params))
            lexical_rows = cursor.fetchall()

            cursor.execute(semantic_query, tuple(semantic_params))
            semantic_rows = cursor.fetchall()

    candidates_map: dict[str, dict[str, Any]] = {}

    for row in lexical_rows:
        node_id_value = str(row[0])
        title = str(row[3])
        code = str(row[4]) if row[4] is not None else None
        lexical_score = score_lexical_node_match(query=query, title=title, code=code)
        if lexical_score <= 0.0:
            continue

        candidates_map[node_id_value] = {
            "node_id": node_id_value,
            "document_id": str(row[1]),
            "node_type": str(row[2]),
            "title": title,
            "code": code,
            "lexical_score": lexical_score,
            "semantic_score": 0.0,
        }

    for row in semantic_rows:
        node_id_value = str(row[0])
        semantic_score = max(0.0, 1.0 - float(row[5]))

        if node_id_value not in candidates_map:
            candidates_map[node_id_value] = {
                "node_id": node_id_value,
                "document_id": str(row[1]),
                "node_type": str(row[2]),
                "title": str(row[3]),
                "code": (str(row[4]) if row[4] is not None else None),
                "lexical_score": 0.0,
                "semantic_score": semantic_score,
            }
        else:
            candidates_map[node_id_value]["semantic_score"] = max(
                float(candidates_map[node_id_value]["semantic_score"]),
                semantic_score,
            )

    ranked: list[NodeCandidate] = []
    for candidate in candidates_map.values():
        lexical_score = float(candidate["lexical_score"])
        semantic_score = float(candidate["semantic_score"])
        final_score = (0.6 * lexical_score) + (0.4 * semantic_score)

        ranked.append(
            NodeCandidate(
                node_id=str(candidate["node_id"]),
                title=str(candidate["title"]),
                code=(str(candidate["code"]) if candidate["code"] is not None else None),
                node_type=str(candidate["node_type"]),
                document_id=str(candidate["document_id"]),
                score=round(final_score, 4),
            )
        )

    ranked.sort(key=lambda item: item.score, reverse=True)
    ranked = ranked[: max(1, min(10, limit))]

    if not ranked:
        return ResolveNodesResponse(
            strong_match=False,
            selected_node_id=None,
            candidates=[],
        )

    top_score = ranked[0].score
    second_score = ranked[1].score if len(ranked) > 1 else 0.0
    strong_match = top_score >= NODE_SCORE_STRONG_THRESHOLD and (
        (top_score - second_score) >= NODE_SCORE_MARGIN_THRESHOLD or top_score >= 0.9
    )

    return ResolveNodesResponse(
        strong_match=strong_match,
        selected_node_id=ranked[0].node_id if strong_match else None,
        candidates=ranked,
    )


def build_node_choice_answer(candidates: list[NodeCandidate]) -> str:
    lines = ["Necesito que elijas el tramite o nodo exacto. Responde solo con el numero:"]
    for index, candidate in enumerate(candidates, start=1):
        suffix = f" ({candidate.code})" if candidate.code else ""
        lines.append(f"{index}. {candidate.title}{suffix}")

    return "\n".join(lines)


def verify_node_alignment(
    results: list[RetrieveResult],
    node_title: str,
    node_code: str | None,
) -> bool:
    if not results:
        return False

    title_tokens = tokenize_for_node(node_title)
    normalized_code = node_code.lower() if node_code else None
    required_hits = max(
        NODE_VERIFICATION_MIN_HITS,
        int(ceil(len(results) * NODE_VERIFICATION_MIN_RATIO)),
    )
    hits = 0

    for result in results:
        content_lower = result.content.lower()

        if normalized_code and normalized_code in content_lower:
            hits += 1
            continue

        if not title_tokens:
            continue

        content_tokens = tokenize_for_node(result.content)
        overlap = len(content_tokens.intersection(title_tokens))
        min_overlap = max(1, int(ceil(len(title_tokens) * 0.2)))
        if overlap >= min_overlap:
            hits += 1

    return hits >= required_hits


def resolve_sensitive_field_priority_rules(
    document_key: str | None = None,
    node_type: str | None = None,
) -> list[tuple[set[str], list[str]]]:
    if document_key:
        normalized_document_key = document_key.strip().lower()
        for marker, rules in SENSITIVE_FIELD_PRIORITY_RULES_BY_DOCUMENT_KEY.items():
            if marker in normalized_document_key:
                return rules

    if node_type:
        normalized_node_type = node_type.strip().lower()
        node_type_rules = SENSITIVE_FIELD_PRIORITY_RULES_BY_NODE_TYPE.get(normalized_node_type)
        if node_type_rules is not None:
            return node_type_rules

    return SENSITIVE_FIELD_PRIORITY_RULES_DEFAULT


def detect_preferred_fields(
    query: str,
    document_key: str | None = None,
    node_type: str | None = None,
) -> list[str]:
    query_lower = query.lower()
    selected: list[str] = []
    seen: set[str] = set()
    rules = resolve_sensitive_field_priority_rules(
        document_key=document_key,
        node_type=node_type,
    )

    for keywords, fields in rules:
        if not any(keyword in query_lower for keyword in keywords):
            continue

        for field in fields:
            if field in seen:
                continue
            seen.add(field)
            selected.append(field)

    return selected


def extract_search_terms(text: str, max_terms: int = 6) -> list[str]:
    tokens = re.findall(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9]{4,}", text.lower())
    unique_tokens = list(dict.fromkeys(tokens))
    return unique_tokens[:max_terms]


def fetch_document_id_by_key(cursor: Any, document_key: str) -> str | None:
    cursor.execute(
        """
        SELECT id
        FROM kb_documents
        WHERE document_key = %s
        LIMIT 1
        """,
        (document_key,),
    )
    row = cursor.fetchone()
    return str(row[0]) if row is not None else None


def resolve_parent_node_id(cursor: Any, document_id: str, dependency_value: str) -> str | None:
    cursor.execute(
        """
        SELECT id
        FROM kb_nodes
        WHERE document_id = %s::uuid
          AND code = %s
        LIMIT 1
        """,
        (document_id, dependency_value),
    )
    row = cursor.fetchone()
    if row is not None:
        return str(row[0])

    cursor.execute(
        """
        SELECT id
        FROM kb_nodes
        WHERE document_id = %s::uuid
          AND lower(title) = lower(%s)
        LIMIT 1
        """,
        (document_id, dependency_value),
    )
    row = cursor.fetchone()
    if row is not None:
        return str(row[0])

    return None


def upsert_excel_node(
    cursor: Any,
    document_id: str,
    node_type: str,
    title: str,
    code: str | None,
    status: str,
    tags: list[str],
) -> str:
    metadata = {"tags": tags} if tags else {}
    metadata_json = json.dumps(metadata, ensure_ascii=False)

    if code:
        cursor.execute(
            """
            SELECT id
            FROM kb_nodes
            WHERE document_id = %s::uuid
              AND node_type = %s
              AND code = %s
            LIMIT 1
            """,
            (document_id, node_type, code),
        )
    else:
        cursor.execute(
            """
            SELECT id
            FROM kb_nodes
            WHERE document_id = %s::uuid
              AND node_type = %s
              AND lower(title) = lower(%s)
            LIMIT 1
            """,
            (document_id, node_type, title),
        )

    existing = cursor.fetchone()
    if existing is not None:
        node_id = str(existing[0])
        cursor.execute(
            """
            UPDATE kb_nodes
            SET title = %s,
                code = %s,
                status = %s,
                metadata = %s::jsonb,
                updated_at = NOW()
            WHERE id = %s::uuid
            """,
            (title, code, status, metadata_json, node_id),
        )
        return node_id

    node_id = str(uuid.uuid4())
    cursor.execute(
        """
        INSERT INTO kb_nodes (
            id, document_id, parent_id, node_type, title, code, status,
            valid_from, valid_to, created_by, updated_by, change_note, metadata,
            created_at, updated_at
        )
        VALUES (
            %s::uuid, %s::uuid, NULL, %s, %s, %s, %s,
            NULL, NULL, NULL, NULL, 'imported from excel', %s::jsonb,
            NOW(), NOW()
        )
        """,
        (node_id, document_id, node_type, title, code, status, metadata_json),
    )
    return node_id


def upsert_excel_chunks_for_node(
    cursor: Any,
    document_id: str,
    node_id: str,
    row: dict[str, Any],
    field_order: list[str],
    pdf_page_start: int | None,
    pdf_page_end: int | None,
    tags: list[str],
) -> tuple[int, list[tuple[str, str]]]:
    cursor.execute(
        """
        SELECT id, metadata->>'field' AS field
        FROM kb_chunks
        WHERE node_id = %s::uuid
          AND COALESCE(metadata->>'source', '') = 'excel'
        """,
        (node_id,),
    )
    existing_chunks = {
        str(db_row[1]): str(db_row[0])
        for db_row in cursor.fetchall()
        if db_row[1] is not None
    }

    desired_fields: set[str] = set()
    upserted = 0
    chunks_to_embed: list[tuple[str, str]] = []

    for field_index, field_name in enumerate(field_order):
        raw_value = row.get(field_name)
        field_value = excel_value_to_str(raw_value)
        if field_value is None:
            continue

        desired_fields.add(field_name)
        metadata_payload: dict[str, Any] = {
            "field": field_name,
            "source": "excel",
        }
        if pdf_page_start is not None:
            metadata_payload["pdf_page_start"] = pdf_page_start
        if pdf_page_end is not None:
            metadata_payload["pdf_page_end"] = pdf_page_end
        if tags:
            metadata_payload["tags"] = tags
        metadata_json = json.dumps(metadata_payload, ensure_ascii=False)

        chunk_id = existing_chunks.get(field_name)
        if chunk_id is not None:
            cursor.execute(
                """
                UPDATE kb_chunks
                SET document_id = %s::uuid,
                    node_id = %s::uuid,
                    chunk_index = %s,
                    content = %s,
                    page = NULL,
                    metadata = %s::jsonb,
                    embedding = NULL
                WHERE id = %s::uuid
                """,
                (document_id, node_id, field_index, field_value, metadata_json, chunk_id),
            )
        else:
            chunk_id = str(uuid.uuid4())
            cursor.execute(
                """
                INSERT INTO kb_chunks (
                    id, document_id, node_id, chunk_index, content, page, metadata, created_at, embedding
                )
                VALUES (
                    %s::uuid, %s::uuid, %s::uuid, %s, %s, NULL, %s::jsonb, NOW(), NULL
                )
                """,
                (chunk_id, document_id, node_id, field_index, field_value, metadata_json),
            )

        upserted += 1
        chunks_to_embed.append((chunk_id, field_value))

    stale_fields = [field for field in existing_chunks.keys() if field not in desired_fields]
    if stale_fields:
        stale_ids = [existing_chunks[field] for field in stale_fields]
        cursor.execute(
            "DELETE FROM kb_chunks WHERE id = ANY(%s::uuid[])",
            (stale_ids,),
        )

    return upserted, chunks_to_embed


def run_ingest_excel(payload: IngestExcelRequest) -> IngestExcelResponse:
    mode = payload.mode.strip().lower()
    if mode != "upsert":
        raise ValueError("mode no soportado. Use 'upsert'.")

    if OPENPYXL_IMPORT_ERROR is not None:
        raise RuntimeError(f"openpyxl no disponible: {OPENPYXL_IMPORT_ERROR}")

    excel_path = validate_ingest_excel_path(payload.path)

    workbook = load_workbook(filename=excel_path, data_only=True)  # type: ignore[misc]
    errors: list[str] = []

    documents_header_error = validate_sheet_headers(workbook, "DOCUMENTOS", DOCUMENT_HEADERS)
    if documents_header_error:
        errors.append(documents_header_error)

    tupa_header_error = validate_optional_sheet_headers(workbook, "TUPA_TRAMITES", TUPA_HEADERS)
    if tupa_header_error:
        errors.append(tupa_header_error)

    rof_header_error = validate_optional_sheet_headers(workbook, "ROF_UNIDADES", ROF_HEADERS)
    if rof_header_error:
        errors.append(rof_header_error)

    documents_rows = read_sheet_rows(workbook, "DOCUMENTOS")
    tupa_rows = read_sheet_rows(workbook, "TUPA_TRAMITES")
    rof_rows = read_sheet_rows(workbook, "ROF_UNIDADES")

    if not tupa_rows and not rof_rows:
        errors.append(
            "Debe incluir datos en al menos una hoja de contenido: TUPA_TRAMITES o ROF_UNIDADES."
        )

    if errors:
        return IngestExcelResponse(
            documents_upserted=0,
            nodes_upserted=0,
            chunks_upserted=0,
            embedded=0,
            errors=errors,
        )

    documents_upserted = 0
    nodes_upserted = 0
    chunks_upserted = 0
    embedded = 0

    if not documents_rows:
        errors.append("Hoja DOCUMENTOS vacia o inexistente.")

    chunks_to_embed_map: dict[str, str] = {}
    pending_parent_links: list[tuple[str, str, str]] = []
    document_key_map: dict[str, str] = {}

    with db_connection() as connection:
        with connection.cursor() as cursor:
            for row in documents_rows:
                row_number = int(row.get("_row_number", 0))
                document_key = excel_value_to_str(row.get("document_key"))
                if document_key is None:
                    errors.append(f"DOCUMENTOS fila {row_number}: document_key es obligatorio.")
                    continue

                title = excel_value_to_str(row.get("title")) or document_key
                doc_type = (excel_value_to_str(row.get("doc_type")) or "txt").lower()
                source_url = excel_value_to_str(row.get("source_url"))
                status = (excel_value_to_str(row.get("status")) or "published").lower()
                version = excel_value_to_str(row.get("version"))

                metadata_payload: dict[str, Any] = {}
                issued_date = excel_value_to_str(row.get("issued_date"))
                entity = excel_value_to_str(row.get("entity"))
                notes = excel_value_to_str(row.get("notes"))
                if issued_date is not None:
                    metadata_payload["issued_date"] = issued_date
                if entity is not None:
                    metadata_payload["entity"] = entity
                if notes is not None:
                    metadata_payload["notes"] = notes
                metadata_json = json.dumps(metadata_payload, ensure_ascii=False)

                existing_doc_id = fetch_document_id_by_key(cursor, document_key)
                if existing_doc_id is not None:
                    cursor.execute(
                        """
                        UPDATE kb_documents
                        SET title = %s,
                            type = %s,
                            source_url = %s,
                            status = %s,
                            version = %s,
                            metadata = %s::jsonb,
                            updated_at = NOW()
                        WHERE id = %s::uuid
                        """,
                        (
                            title,
                            doc_type,
                            source_url,
                            status,
                            version,
                            metadata_json,
                            existing_doc_id,
                        ),
                    )
                    document_key_map[document_key] = existing_doc_id
                else:
                    document_id = str(uuid.uuid4())
                    cursor.execute(
                        """
                        INSERT INTO kb_documents (
                            id, document_key, title, type, source_url, status, version, metadata, created_at, updated_at
                        )
                        VALUES (
                            %s::uuid, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW(), NOW()
                        )
                        """,
                        (
                            document_id,
                            document_key,
                            title,
                            doc_type,
                            source_url,
                            status,
                            version,
                            metadata_json,
                        ),
                    )
                    document_key_map[document_key] = document_id

                documents_upserted += 1

            for row in tupa_rows:
                row_number = int(row.get("_row_number", 0))
                document_key = excel_value_to_str(row.get("document_key"))
                if document_key is None:
                    errors.append(f"TUPA_TRAMITES fila {row_number}: document_key es obligatorio.")
                    continue

                document_id = document_key_map.get(document_key) or fetch_document_id_by_key(cursor, document_key)
                if document_id is None:
                    errors.append(
                        f"TUPA_TRAMITES fila {row_number}: document_key '{document_key}' no encontrado."
                    )
                    continue

                title = excel_value_to_str(row.get("procedure_title"))
                if title is None:
                    errors.append(f"TUPA_TRAMITES fila {row_number}: procedure_title es obligatorio.")
                    continue

                code = excel_value_to_str(row.get("procedure_code"))
                status = (excel_value_to_str(row.get("status")) or "published").lower()
                tags = parse_tags(row.get("tags"))
                node_id = upsert_excel_node(
                    cursor=cursor,
                    document_id=document_id,
                    node_type="procedure",
                    title=title,
                    code=code,
                    status=status,
                    tags=tags,
                )
                nodes_upserted += 1

                pdf_page_start = excel_value_to_int(row.get("pdf_page_start"))
                pdf_page_end = excel_value_to_int(row.get("pdf_page_end"))
                upserted_count, embed_items = upsert_excel_chunks_for_node(
                    cursor=cursor,
                    document_id=document_id,
                    node_id=node_id,
                    row=row,
                    field_order=PROCEDURE_CHUNK_FIELDS,
                    pdf_page_start=pdf_page_start,
                    pdf_page_end=pdf_page_end,
                    tags=tags,
                )
                chunks_upserted += upserted_count
                for chunk_id, content in embed_items:
                    chunks_to_embed_map[chunk_id] = content

            for row in rof_rows:
                row_number = int(row.get("_row_number", 0))
                document_key = excel_value_to_str(row.get("document_key"))
                if document_key is None:
                    errors.append(f"ROF_UNIDADES fila {row_number}: document_key es obligatorio.")
                    continue

                document_id = document_key_map.get(document_key) or fetch_document_id_by_key(cursor, document_key)
                if document_id is None:
                    errors.append(
                        f"ROF_UNIDADES fila {row_number}: document_key '{document_key}' no encontrado."
                    )
                    continue

                title = excel_value_to_str(row.get("org_title"))
                if title is None:
                    errors.append(f"ROF_UNIDADES fila {row_number}: org_title es obligatorio.")
                    continue

                code = excel_value_to_str(row.get("org_code"))
                status = (excel_value_to_str(row.get("status")) or "published").lower()
                tags = parse_tags(row.get("tags"))
                node_id = upsert_excel_node(
                    cursor=cursor,
                    document_id=document_id,
                    node_type="org_unit",
                    title=title,
                    code=code,
                    status=status,
                    tags=tags,
                )
                nodes_upserted += 1

                dependency = excel_value_to_str(row.get("dependencia_superior"))
                if dependency:
                    pending_parent_links.append((node_id, document_id, dependency))

                pdf_page_start = excel_value_to_int(row.get("pdf_page_start"))
                pdf_page_end = excel_value_to_int(row.get("pdf_page_end"))
                upserted_count, embed_items = upsert_excel_chunks_for_node(
                    cursor=cursor,
                    document_id=document_id,
                    node_id=node_id,
                    row=row,
                    field_order=ORG_UNIT_CHUNK_FIELDS,
                    pdf_page_start=pdf_page_start,
                    pdf_page_end=pdf_page_end,
                    tags=tags,
                )
                chunks_upserted += upserted_count
                for chunk_id, content in embed_items:
                    chunks_to_embed_map[chunk_id] = content

            for node_id, document_id, dependency_value in pending_parent_links:
                parent_id = resolve_parent_node_id(cursor, document_id, dependency_value)
                if parent_id is None or parent_id == node_id:
                    continue
                cursor.execute(
                    """
                    UPDATE kb_nodes
                    SET parent_id = %s::uuid,
                        updated_at = NOW()
                    WHERE id = %s::uuid
                    """,
                    (parent_id, node_id),
                )

            if payload.embed and chunks_to_embed_map:
                chunk_ids = list(chunks_to_embed_map.keys())
                chunk_texts = [chunks_to_embed_map[chunk_id] for chunk_id in chunk_ids]
                model = get_embedding_model()
                embeddings = model.encode(
                    chunk_texts,
                    convert_to_numpy=True,
                    show_progress_bar=False,
                    normalize_embeddings=True,
                )
                params: list[tuple[str, str]] = []
                for chunk_id, embedding in zip(chunk_ids, embeddings):
                    values = embedding.tolist()
                    if len(values) != EMBEDDING_DIMENSION:
                        raise ValueError(
                            f"Invalid embedding dimension for chunk {chunk_id}: "
                            f"{len(values)} (expected {EMBEDDING_DIMENSION})"
                        )
                    params.append((to_vector_literal(values), chunk_id))

                execute_batch(
                    cursor,
                    "UPDATE kb_chunks SET embedding = %s::vector WHERE id = %s::uuid",
                    params,
                    page_size=100,
                )
                embedded = len(params)

            connection.commit()

    return IngestExcelResponse(
        documents_upserted=documents_upserted,
        nodes_upserted=nodes_upserted,
        chunks_upserted=chunks_upserted,
        embedded=embedded,
        errors=errors,
    )


def build_rag_prompt(message: str, results: list[RetrieveResult]) -> str:
    context_blocks = []
    for index, result in enumerate(results, start=1):
        context_blocks.append(
            "\n".join(
                [
                    f"[{index}] title: {result.document.title}",
                    f"chunk_id: {result.chunk_id}",
                    f"type: {result.document.type}",
                    f"content: {result.content}",
                ]
            )
        )

    context = "\n\n---\n\n".join(context_blocks)

    return "\n".join(
        [
            "Sistema:",
            (
                "Eres un asistente institucional. Responde SOLO usando el CONTEXTO. "
                "No inventes plazos, montos, requisitos ni procedimientos. "
                "Si no esta explicito, responde: "
                "'No encuentro sustento en la base cargada para responder eso.' "
                "Responde breve y clara."
            ),
            "",
            "Usuario:",
            message,
            "",
            "Contexto:",
            context,
        ]
    )


def build_context_text(results: list[RetrieveResult]) -> str:
    return "\n".join(result.content for result in results if result.content.strip() != "")


def split_sentences(text: str) -> list[str]:
    parts = re.split(r"(?<=[.!?])\s+", text.strip())
    return [part.strip() for part in parts if part.strip() != ""]


def sentence_has_unsupported_sensitive_data(sentence: str, context_text: str) -> bool:
    for pattern in SENSITIVE_DATA_PATTERNS:
        if pattern.search(sentence) and not pattern.search(context_text):
            return True

    return False


def apply_anti_hallucination_guardrail(answer: str, context_text: str) -> str:
    cleaned_answer = re.sub(r"\s+", " ", answer).strip()
    if cleaned_answer == "":
        return NO_SUPPORT_ANSWER

    if context_text.strip() == "":
        return NO_SUPPORT_ANSWER

    sentences = split_sentences(cleaned_answer)
    if not sentences:
        return NO_SUPPORT_ANSWER

    safe_sentences = [
        sentence
        for sentence in sentences
        if not sentence_has_unsupported_sensitive_data(sentence, context_text)
    ]

    if not safe_sentences:
        return NO_SUPPORT_ANSWER

    return " ".join(safe_sentences).strip()


def dedupe_sources(sources: list[ChatSource], max_items: int = 3) -> list[ChatSource]:
    seen: set[str] = set()
    deduped: list[ChatSource] = []

    for source in sources:
        normalized_title = source.title.strip().lower()
        normalized_url = (source.url or "").strip().lower()
        key = normalized_url if normalized_url != "" else f"{normalized_title}|{normalized_url}"

        if key in seen:
            continue

        seen.add(key)
        deduped.append(source)

        if len(deduped) >= max_items:
            break

    return deduped


def generate_answer_with_llm(message: str, results: list[RetrieveResult]) -> str:
    if not results:
        return NO_SUPPORT_ANSWER

    provider = get_llm_provider()
    prompt = build_rag_prompt(message, results)
    return provider.generate(
        prompt=prompt,
        max_tokens=LLM_MAX_TOKENS,
        temperature=LLM_TEMPERATURE,
    )


def reindex_document(document_id: str, batch_size: int = REINDEX_BATCH_SIZE) -> int:
    processed = 0
    model = get_embedding_model()

    with db_connection() as connection:
        with connection.cursor() as cursor:
            while True:
                cursor.execute(
                    """
                    SELECT id, content
                    FROM kb_chunks
                    WHERE document_id = %s::uuid
                      AND embedding IS NULL
                    ORDER BY chunk_index, id
                    LIMIT %s
                    """,
                    (document_id, batch_size),
                )
                rows = cursor.fetchall()

                if not rows:
                    break

                chunk_ids = [str(row[0]) for row in rows]
                texts = [str(row[1] or "") for row in rows]

                embeddings = model.encode(
                    texts,
                    convert_to_numpy=True,
                    show_progress_bar=False,
                    normalize_embeddings=True,
                )

                params: list[tuple[str, str]] = []
                for chunk_id, embedding in zip(chunk_ids, embeddings):
                    embedding_list = embedding.tolist()
                    if len(embedding_list) != EMBEDDING_DIMENSION:
                        raise ValueError(
                            f"Invalid embedding dimension for chunk {chunk_id}: "
                            f"{len(embedding_list)} (expected {EMBEDDING_DIMENSION})"
                        )

                    params.append((to_vector_literal(embedding_list), chunk_id))

                execute_batch(
                    cursor,
                    "UPDATE kb_chunks SET embedding = %s::vector WHERE id = %s::uuid",
                    params,
                    page_size=batch_size,
                )
                connection.commit()
                processed += len(params)

    return processed


def run_reindex_node(
    node_id: str,
    batch_size: int = REINDEX_BATCH_SIZE,
    force: bool = True,
) -> ReindexNodeResponse:
    processed = 0
    model = get_embedding_model()

    with db_connection() as connection:
        with connection.cursor() as cursor:
            if force:
                cursor.execute(
                    "UPDATE kb_chunks SET embedding = NULL WHERE node_id = %s::uuid",
                    (node_id,),
                )
                connection.commit()

            while True:
                cursor.execute(
                    """
                    SELECT id, content
                    FROM kb_chunks
                    WHERE node_id = %s::uuid
                      AND embedding IS NULL
                    ORDER BY chunk_index, id
                    LIMIT %s
                    """,
                    (node_id, batch_size),
                )
                rows = cursor.fetchall()

                if not rows:
                    break

                chunk_ids = [str(row[0]) for row in rows]
                texts = [str(row[1] or "") for row in rows]

                embeddings = model.encode(
                    texts,
                    convert_to_numpy=True,
                    show_progress_bar=False,
                    normalize_embeddings=True,
                )

                params: list[tuple[str, str]] = []
                for chunk_id, embedding in zip(chunk_ids, embeddings):
                    embedding_list = embedding.tolist()
                    if len(embedding_list) != EMBEDDING_DIMENSION:
                        raise ValueError(
                            f"Invalid embedding dimension for chunk {chunk_id}: "
                            f"{len(embedding_list)} (expected {EMBEDDING_DIMENSION})"
                        )

                    params.append((to_vector_literal(embedding_list), chunk_id))

                execute_batch(
                    cursor,
                    "UPDATE kb_chunks SET embedding = %s::vector WHERE id = %s::uuid",
                    params,
                    page_size=batch_size,
                )
                connection.commit()
                processed += len(params)

            cursor.execute(
                """
                SELECT COUNT(*)
                FROM kb_chunks
                WHERE node_id = %s::uuid
                  AND embedding IS NULL
                """,
                (node_id,),
            )
            remaining = int(cursor.fetchone()[0])

    return ReindexNodeResponse(
        node_id=node_id,
        processed=processed,
        remaining=remaining,
    )


def run_reindex(batch_size: int = REINDEX_BATCH_SIZE) -> ReindexResponse:
    processed = 0

    model = get_embedding_model()

    with db_connection() as connection:
        with connection.cursor() as cursor:
            while True:
                cursor.execute(
                    """
                    SELECT id, content
                    FROM kb_chunks
                    WHERE embedding IS NULL
                    ORDER BY created_at, id
                    LIMIT %s
                    """,
                    (batch_size,),
                )
                rows = cursor.fetchall()

                if not rows:
                    break

                chunk_ids = [str(row[0]) for row in rows]
                texts = [str(row[1] or "") for row in rows]

                embeddings = model.encode(
                    texts,
                    convert_to_numpy=True,
                    show_progress_bar=False,
                    normalize_embeddings=True,
                )

                params: list[tuple[str, str]] = []

                for chunk_id, embedding in zip(chunk_ids, embeddings):
                    embedding_list = embedding.tolist()

                    if len(embedding_list) != EMBEDDING_DIMENSION:
                        raise ValueError(
                            f"Invalid embedding dimension for chunk {chunk_id}: "
                            f"{len(embedding_list)} (expected {EMBEDDING_DIMENSION})"
                        )

                    params.append((to_vector_literal(embedding_list), chunk_id))

                execute_batch(
                    cursor,
                    "UPDATE kb_chunks SET embedding = %s::vector WHERE id = %s::uuid",
                    params,
                    page_size=batch_size,
                )
                connection.commit()
                processed += len(params)

            cursor.execute("SELECT COUNT(*) FROM kb_chunks WHERE embedding IS NULL")
            remaining = int(cursor.fetchone()[0])

    return ReindexResponse(processed=processed, remaining=remaining)


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/retrieve", response_model=RetrieveResponse)
def retrieve(payload: RetrieveRequest) -> RetrieveResponse:
    try:
        results = retrieve_semantic(payload.query, payload.top_k, payload.filters)
        return RetrieveResponse(results=results)
    except Exception as exception:
        raise HTTPException(status_code=500, detail=f"retrieve_error: {exception}") from exception


@app.post("/resolve/nodes", response_model=ResolveNodesResponse)
def resolve_nodes(payload: ResolveNodesRequest) -> ResolveNodesResponse:
    try:
        return resolve_nodes_hybrid(
            query=payload.query,
            document_id=payload.document_id,
            node_type=payload.node_type,
            limit=payload.limit,
        )
    except Exception as exception:
        raise HTTPException(status_code=500, detail=f"resolve_nodes_error: {exception}") from exception


@app.post("/chat", response_model=ChatResponse)
async def chat(payload: ChatRequest) -> ChatResponse:
    def run_chat() -> ChatResponse:
        selected_node: dict[str, str | None] | None = None
        resolved: ResolveNodesResponse | None = None

        if payload.selected_node_id:
            selected_node = fetch_published_node(payload.selected_node_id)
            if selected_node is None:
                return ChatResponse(answer=NO_SUPPORT_ANSWER, sources=[])
        else:
            resolved = resolve_nodes_hybrid(
                query=payload.message,
                document_id=payload.document_id,
                limit=5,
            )
            if not resolved.candidates:
                return ChatResponse(answer=NO_SUPPORT_ANSWER, sources=[])

            selected_node_id: str | None = None
            if resolved.strong_match and resolved.selected_node_id is not None:
                selected_node_id = resolved.selected_node_id
            elif len(resolved.candidates) == 1:
                selected_node_id = resolved.candidates[0].node_id

            if selected_node_id is None:
                return ChatResponse(
                    answer=build_node_choice_answer(resolved.candidates[:5]),
                    sources=[],
                )

            selected_node = fetch_published_node(selected_node_id)
            if selected_node is None:
                return ChatResponse(answer=NO_SUPPORT_ANSWER, sources=[])

        node_id = str(selected_node["id"])
        document_id = (
            str(selected_node["document_id"])
            if selected_node.get("document_id") is not None
            else payload.document_id
        )
        preferred_fields = detect_preferred_fields(
            payload.message,
            document_key=(
                str(selected_node["document_key"])
                if selected_node.get("document_key") is not None
                else None
            ),
            node_type=(
                str(selected_node["node_type"])
                if selected_node.get("node_type") is not None
                else None
            ),
        )
        results = retrieve_semantic(
            query=payload.message,
            top_k=payload.top_k,
            filters=RetrieveFilters(
                document_id=document_id,
                node_id=node_id,
            ),
            preferred_fields=preferred_fields,
        )

        if not results:
            return ChatResponse(answer=NO_SUPPORT_ANSWER, sources=[])

        sources = [
            ChatSource(
                title=result.document.title,
                type=result.document.type,
                reference=f"chunk:{result.chunk_id}",
                url=result.document.url,
            )
            for result in results
        ]
        sources = dedupe_sources(sources, max_items=3)

        try:
            raw_answer = generate_answer_with_llm(payload.message, results)
            context_text = build_context_text(results)
            answer = apply_anti_hallucination_guardrail(raw_answer, context_text)
            if answer == NO_SUPPORT_ANSWER:
                return ChatResponse(answer=answer, sources=[])

            return ChatResponse(answer=answer, sources=sources)
        except Exception as generation_exception:
            message = str(generation_exception)
            if "ollama" in message.lower():
                answer = (
                    "El servicio no disponible: Ollama no responde. "
                    "Verifique que este activo e intente nuevamente."
                )
            else:
                answer = "El servicio no disponible. Intente nuevamente mas tarde."

            return ChatResponse(answer=answer, sources=[])

    try:
        return await run_in_threadpool(run_chat)
    except Exception as exception:
        raise HTTPException(status_code=500, detail=f"chat_error: {exception}") from exception


@app.post("/ingest/excel", response_model=IngestExcelResponse)
async def ingest_excel(payload: IngestExcelRequest) -> IngestExcelResponse:
    try:
        return await run_in_threadpool(run_ingest_excel, payload)
    except ValueError as validation_exception:
        raise HTTPException(status_code=422, detail=str(validation_exception)) from validation_exception
    except Exception as exception:
        raise HTTPException(status_code=500, detail=f"ingest_excel_error: {exception}") from exception


@app.post("/reindex/node", response_model=ReindexNodeResponse)
async def reindex_node(payload: ReindexNodeRequest) -> ReindexNodeResponse:
    try:
        return await run_in_threadpool(
            run_reindex_node,
            payload.node_id,
            REINDEX_BATCH_SIZE,
            payload.force,
        )
    except Exception as exception:
        raise HTTPException(status_code=500, detail=f"reindex_node_error: {exception}") from exception


@app.post("/reindex", response_model=ReindexResponse)
async def reindex() -> ReindexResponse:
    try:
        return await run_in_threadpool(run_reindex, REINDEX_BATCH_SIZE)
    except Exception as exception:
        raise HTTPException(status_code=500, detail=f"reindex_error: {exception}") from exception
